# First exercise
# name = 'John Smith'
# age = 20
# is_new = True
#
# name = input('What is your name? ')
# print('Hi ' + name)
# favorite_color = input('What is your favorite color? ')
# print(name + ' likes '+ favorite_color)

# Second Exersice
# birth_year = input('Birth year: ')
# age = 2019 - int(birth_year)
# print(age)

# Third exercise
# weight_lbs = input('Weight (lbs): ')
# weight_kg = float(weight_lbs) * 0.45
# print("You are " + str(weight_kg) + ' kg')

# Forth exercise

# course = "Python's Course For Beginners"
# email_message = ''''
# Hi Eric,
#
# Here is the first mail I send you through Python
#
# Thank you,
# Paco
#
# '''
#
# print(email_message)
# #First letter
# print(course[1])
#
# #Last letter
# print(course[-1])
#
# #Range
# print(course[0:3])
#
# first = 'John'
# last = 'Doe'
# message = first + ' [' + last + '] ' + 'is a coder'
# #String format
# msg = f'{first} [{last}] is a coder'
# print(msg)
#
# print(len(course))
#
# #String functions
# print(course.upper())
#
# print(course.replace('Beginners','Absolute Beginners'))
# #Exist
# print('Python' in course)


# Fifth exercise
#
# import math
#
# x = 2.9
# print(round(x))
# print(abs(x))
# print(math.ceil(x))
#
# price = 1000000
# has_good_credit = True
# has_high_income = True
#
# if has_good_credit and has_high_income:
# 	payment = price * 0.10
# else:
# 	payment = price * 0.20
# print(f'Payment: {payment}$')
#
# has_criminal_record = False
#
# if has_good_credit and not has_criminal_record:
# 	print("He's clean!")
#
# name = input("Who's your name? ")
# if len(name) < 3:
# 	print("Name must be at least 3 characters")
# elif len(name) > 50:
# 	print("Name can be a maximum of 50 characters")
# else:
# 	print("Your name looks good!")

# Sixth exercise
# weight = float(input('Weight: '))
# unit = input('Lbs(L) or Kg(K): ')
# if unit.upper() == 'L':
# 	converted = weight * 0.45
# 	print(f'Converted weight: {converted} pounds')
# elif unit.upper() == 'K':
# 	converted = weight / 0.45
# 	print(f'Converted weight: {converted} pounds')
# else:
# 	print("Incorrect unit")

# Seventh exercise - Loops
# i = 1
# while i <= 5:
# 	print('*' * i)
# 	i += 1
#
# # Game 1 - Guess the secret number
# secret_number  = 9
# opportunities = 3
# while opportunities > 0:
# 	opportunities -= 1
# 	guess = int(input("Guess: "))
# 	if guess == secret_number:
# 		print("You won")
# 		break
# else:
# 	print("You failed!")


# Game 1 - Car game
# command = ""
# started = False
# while True:
# 	command = input("> ").lower()
# 	if command == "start":
# 		if started:
# 			print("Car is alredy started!")
# 		else:
# 			started = True
# 			print("Car started...")
# 	elif command == "stop":
# 		if not started:
# 			print("Car is alredy stopped!")
# 		else:
# 			started = False
# 			print("Car stopped")
# 	elif command == "help":
# 		print('''
# start - to start the car
# stop - to stop the car
# quit - to exit
# ''')
# 	elif command == "quit":
# 		print("Bye bye!")
# 		break
# 	else:
# 		print("Incorrect command")

# Eighth exercise - For loop

# for char in 'Python':
# 	print(char)
#
# for item in ['Alex','Lidia','Paula']:
# 	print(item)
#
# for n in range(5, 10):
# 	print(n)
#
# prices = [30,40,50,60]
# total = 0
# for price in prices:
# 	total += price
# print(f'Total: {total}')
#
# for x in range(4):
# 	for y in range(3):
# 		print(f'({x},{y})')

# numbers = [5, 2, 5, 2, 2]
# for number in numbers:
# 	print('x' * number)
#
# for number in numbers:
# 	output = ''
# 	for x in range(number):
# 		output += 'x'
# 	print(output)

# Nineth exercise
# names = ['Lidia','Eric','Josefa','Epi']
# print(names[1])
# numbers = [4, 5, 1, 3, 6, 1, 2]
# max = numbers[0]
# for number in numbers:
# 	if number > max:
# 		max=number
# print(max)
#
# matrix = [
# 	[1, 2, 3],
# 	[4, 5, 6],
# 	[7, 8, 9]
# ]
# matrix[0][0] = 59
# print(matrix[0][0])
#
# for row in matrix:
# 	for item in row:
# 		print(item)
# numbers.sort()
# numbers.reverse()
# numbers2 = numbers.copy()
# print(numbers)
# numbers.append(20)
# numbers.insert(0,10)
# print(numbers.index(5))
# numbers.remove(5)
# numbers.pop()
# numbers2.clear()
# print(numbers)
#
# #Game 2 - Find duplicates
# list = []
# for number in numbers:
# 	if number not in list:
# 		list.append(number)
# print(list)

# Exercise Tenth - Tuples and unpacking
# numbers = (1, 2, 3)
# print(numbers[0])
#
# coordinates = (1, 2, 3)
# coordinates2 = [1, 2, 3]
#
# x =  coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
#
# x1, y1, z1 = coordinates
#
# x2, y2, z2 = coordinates2
#
# print(x1)
# print(x2)
#
# #Eleventh exercise - Dictionaries
#
# customer = {
# 	"name": "John Doe",
# 	"age": 20,
# 	"is_verified": True
# }
#
# customer["name"] = "Jack Doe"
# customer["birthday"] = "11-11-1996"
#
# print(customer["name"])
# print(customer.get("name"))
# print(customer.get("birthday"))
# numbers = {
# 	1: "One",
# 	2: "Two",
# 	3: "Three",
# 	4: "Four",
# 	5: "Five",
# 	6: "Six",
# 	7: "Seven",
# 	8: "Eight",
# 	9: "Nine"
# }
# phone = input("Phone: ")
# output = ""
# for number in phone:
# 	output += numbers.get(int(number)) + ' '
# print(output)
#
# emojis = {
# 	":)": "😊",
# 	";)": "😉",
# 	":(": "😢",
# 	":d": "😁",
# 	"xd": "😆",
# 	":p": "😜"
# }
#
# while True:
# 	msg = input("> ")
# 	words = msg.split(' ')
# 	output = ""
# 	if msg == "quit":
# 		print("Bye bye")
# 		break
# 	for word in words:
# 		output += emojis.get(word.lower(),word) + ' '
# 	print(output)

# Twelve exercise - Functions
# def greet_user(name,last_name):
# 	print(f"Hi there {name} {last_name}")
# 	print("Welcome aboard")
# #Positional argument
# greet_user("Eric","Nunez")
#
# #Keyword argument
# greet_user(last_name="Nunez",name="Eric")
#
# #Mixed
# greet_user("Eric",last_name="Nunez")
#
# def square(number):
# 	return number * number
#
# print(square(2))
#
# def emojis_converter(msg):
# 	emojis = {
# 		":)": "😊",
# 		";)": "😉",
# 		":(": "😢",
# 		":d": "😁",
# 		"xd": "😆",
# 		":p": "😜"
# 	}
#
# 	words = msg.split(' ')
# 	output = ""
# 	for word in words:
# 		output += emojis.get(word.lower(),word) + ' '
# 	return output
#
# msg = input("> ")
# print(emojis_converter(msg))

# Thirdteenth - Exceptions
# try:
# 	age = int(input("Age: "))
# 	income = 20000
# 	risk = income / age
# 	print(age)
# except ZeroDivisionError:
# 	print("Age cannot be zero!")
# except ValueError:
# 	print("Invalid value")

# Fourteenth exercise - Classes
#
# class Point:
# 	#Constructor
# 	def __init__(self,x,y):
# 		self.x = x
# 		self.y = y
#
# 	def move(self):
# 		print("move")
# 	def draw(self):
# 		print("draw")
#
#
# point1 = Point(10,20)
# print(point1.x)
#
# class Person:
# 	def __init__(self,name):
# 		self.name = name
#
# 	def talk(self):
# 		print(f"Hi, I'm {self.name}")
#
#
# person = Person("Bob")
# print(person.name)
# person.talk()


# class Mammal:
# 	def walk(self):
# 		print("walk")
#
# class Dog(Mammal):
# 	def bark(self):
# 		print("bark")
#
# class Cat(Mammal):
# 	def be_annoying(self):
# 		print("annoying")
#
# dog1 = Dog()
# dog1.walk()

# Fifteenth exercise - Modules and packages
# Python 3 module index
# import converters, utils
# from converters import lbs_to_kg
# import ecommerce.shipping
# from ecommerce.shipping import calc_shipping
# from ecommerce import shipping
#
# print(converters.kg_to_lbs(70))
#
# print(utils.find_max([3,4,5,6,9,10,56,12,45,67]))
#
# shipping.calc_shipping()

# import random
#
# # for i in range(3):
# # 	print(random.randint(10, 20))
#
# members = ["John", "Mary", "Bob", "Charles"]
# leader = random.choice(members)
# print(leader)
#
#
# class Dice:
# 	def roll(self):
# 		x = random.randint(1, 6)
# 		y = random.randint(1, 6)
# 		return x, y
#
#
# dice = Dice()
# print(dice.roll())

# from pathlib import Path
# #Absolute path
# #c:\Program Files\Microsoft
# # /usr/local/bin
# # Relative path
#
# path = Path()
# # path.mkdir()
# # path.rmdir()
# # print(path.exists())
# for file in path.glob('*.py'):
# 	print(file)

#Sixteenth exercise - Working with pip
#openpyxl - Lib for working with Excel

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
	wb = xl.load_workbook(filename)
	sh1 = wb['Sheet1']

	for row in range(2, sh1.max_row + 1):
		cell = sh1.cell(row, 3)
		corrected_price = cell.value * 0.9
		corrected_price_cell = sh1.cell(row, 4)
		corrected_price_cell.value = corrected_price

	values = Reference(sh1, min_row=2,
					   max_row=sh1.max_row,
					   min_col=4,
					   max_col=4)

	chart = BarChart()
	chart.add_data(values)
	sh1.add_chart(chart, 'e2')

	wb.save(filename)

process_workbook('transactions.xlsx')


